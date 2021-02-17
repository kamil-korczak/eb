#ebapp\classes\report.py
from django.http import HttpResponse
from datetime import datetime
from django.utils import timezone

from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ebapp.models import Auctions, AuctionsDetails, Settings

""" Generate report Excel file """

class Report:

    def __init__(self, account_name, promotion_date_start, settings):

        self.wb = Workbook()

        self.now = datetime.now()
        self.date_string = self.now.strftime("%Y-%m-%d")
        self.time_string = self.now.strftime("%H:%m:%S")

        self.promotion_date_start = promotion_date_start.strftime('%d/%m/%Y')

        self.dest_filename = f'ebay-wow-proposal-report-{account_name}_{self.date_string}_{self.time_string}.xlsx'#.format(self.date_string,self.time_string)
        self.setSettings(settings)
        
    def auctionFormat(self, auction_format_choice):
        auction_format_choice = AuctionsDetails.AUCTION_WEEKLY if auction_format_choice == None else auction_format_choice

        return AuctionsDetails.AUCTION_FORMAT_CHOICES[auction_format_choice][1]

    def isBothAuctionFormat(self, auction_format_choices):
        return True if auction_format_choices == AuctionsDetails.AUCTION_BOTH else False

    def setRowAuctionFormat(self, auction_format_choice, i_row):
        if auction_format_choice == AuctionsDetails.AUCTION_BOTH:
            if i_row == 1:
                return AuctionsDetails.AUCTION_WEEKLY
            else:
                return AuctionsDetails.AUCTION_DAILY
        else:
            return auction_format_choice

    def setRowAuctionData(self, object_row, object_auction_format):
        row = [
            object_row.auction.external_id,
            object_row.auction.title_auction,
            '', #EAN
            object_row.items_stock,
            object_row.price_normal,
            'uvp', #uvp
            object_row.price_uvp,
            self.auctionFormat(object_auction_format),#'weekly/daily'
            self.promotion_date_start,
            '', #Preisvergleichsportal 1 (Amazon für DE)
            '', #Preisvergleichsportal 2 (Idealo für DE)
            object_row.auction.ebay_category, # ebay category
            object_row.auction.seller_comment,
            self.getSettings().seller_email,
            self.getSettings().target_website_for_wow_offer,
        ]

        return row

    def setSettings(self, settings):
        self.settings = settings

    def getSettings(self):
        return self.settings

    def updateDownloadDate(self, auctions_queryset):

        auctions_list = list()

        for auction in auctions_queryset:
            auctions_list.append(auction.auction_id)

        import pytz

        date_str = str(timezone.now())

        date_update = Auctions.objects.filter(id__in=auctions_list).update(time_downloaded=date_str)

    def generate(self, auctions_queryset):
        ws1 = self.wb.active
        ws1.title = "Raport - {}".format(self.date_string) 

        # u" - unicode string

        row_num = 0

        columns = [
            (u"eBay-Artikelnr", 15), #external_id
            (u"Artikelbezeichnung des eBay-Angebots", 15), #title_auction
            (u"EAN", 15), #
            (u"Artikelstückzahl für WOW! Angebot", 15), #stock
            (u"Angebotspreis des Verkäufers", 15), #price_normal
            (u"BISHER-Preis (UK)", 15), #uvp
            (u"BISHER-Preis (UK)_1", 15), #price_uvp
            (u"Format des WOW! Angebots", 15), #weekly/daily
            (u"Artikel frühestens verfügbar für WOW! Angebot ab", 15), # promotions start from
            (u"Preisvergleichsportal 1 (Amazon für DE)", 15), # 
            (u"Preisvergleichsportal 2 (Idealo für DE)", 15), # 
            (u"eBay-Hauptkategorie", 15), # ebay - category
            (u"Kommentare des Verkäufers", 15), # seller's comment
            (u"E-Mail des Verkäufers", 15), # seller's email
            (u"Ziel-Website für WOW! Angebot", 15), # website for wow
        ]

        for col_num in range(len(columns)):
            c = ws1.cell(row=row_num +1, column=col_num+1)
            c.value = columns[col_num][0]
            ws1.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for auction_object in auctions_queryset:
            row_num += 1
            
            if self.isBothAuctionFormat(auction_object.auction_format):
                auction_rows_count = 2
            else:
                auction_rows_count = 1

            for i_row in range(auction_rows_count):
                row = self.setRowAuctionData(
                    auction_object, 
                    self.setRowAuctionFormat(
                            auction_object.auction_format,
                            i_row
                        )
                    )

                for col_num in range(len(row)):
                    c = ws1.cell(row=row_num+1, column=col_num+1)
                    c.value = row[col_num]

                row_num += 1 if auction_rows_count == 2 and i_row == 0 else 0
        
        #save as stream
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename={self.dest_filename}'

        self.updateDownloadDate(auctions_queryset)

        return response
            